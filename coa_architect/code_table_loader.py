"""
code_table_loader.py — Scans the 1.code_tables/ folder and loads advisory reference files.

Advisory files are keyed by column name (e.g., 'FERC Code.pdf' maps to the 'FERC Code'
column).  They provide richer descriptions that augment, but never override, the
authoritative code lists embedded in the CoA workbook tabs.

Supported formats: CSV, Excel (.xlsx / .xls), PDF (text-based only).

Matching rule: filename stem must match an actual column header from the CoA main sheet,
case-insensitively and character-for-character (spaces allowed; hyphens are not interchangeable
with spaces).  For example:
  'ferc code.pdf'  → matches 'FERC Code'   ✓
  'ferc-code.pdf'  → does NOT match        ✗
"""

import csv
import os
import re

import openpyxl


# ---------------------------------------------------------------------------
# Internal field names that have advisory logic implemented in suggester.py
# Columns not in this set are loaded but flagged as having no advisory effect.
# ---------------------------------------------------------------------------
ADVISORY_LOGIC_FIELDS = {
    "ferc_code",
    "asset_life",
    "cash_flow_category",
    "book_tax_difference",
    "posting_edit",
}

# ---------------------------------------------------------------------------
# PDF regex patterns — copied from tools/extract_ferc_pdf.py
# (tools/ is not a package so we cannot import from it at runtime)
# ---------------------------------------------------------------------------

# Matches lines like:  "101   Electric plant in service"
TABULAR_RE = re.compile(
    r"^\s*(\d{2,4}(?:\.\d{1,2})?)\s{2,}(.+)$"
)

# Matches lines like:  "Account 101. Electric Plant in Service."
HEADER_RE = re.compile(
    r"(?i)\bAccount\s+(?:No\.?\s*)?(\d{2,4}(?:\.\d{1,2})?)[.\s\-–]+(.+)"
)

# Matches lines that are purely a code (description may wrap to the next line)
CODE_ONLY_RE = re.compile(
    r"^\s*(\d{2,4}(?:\.\d{1,2})?)\s*$"
)


def _clean_description(text: str) -> str:
    """
    Strips trailing punctuation artifacts and extra whitespace from text
    extracted by pdfplumber (which sometimes introduces extra spaces).
    """
    # Collapse runs of whitespace to a single space
    text = re.sub(r"\s+", " ", text).strip()
    # Remove trailing period that looks like a section terminator
    text = text.rstrip(".")
    return text


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_actual_column_headers(worksheet) -> list:
    """
    Reads row 1 of the given worksheet and returns a list of raw header strings
    (whitespace stripped, non-empty cells only).

    These are the strings that file stems in 1.code_tables/ are matched against.
    """
    headers = []
    for cell in worksheet[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())
    return headers


def scan_code_tables(folder_path: str, column_headers: list) -> tuple:
    """
    Scans folder_path for advisory reference files and matches them to column headers.

    Matching is case-insensitive on the filename stem vs each header string.
    Supported extensions: .csv, .xlsx, .xls, .pdf

    Returns a 3-tuple:
      matches      — {column_header: file_path}  — exactly one file per column
      conflicts    — {column_header: [file_path, ...]}  — two or more files matched
      unrecognized — [file_path, ...]  — no column matched
    """
    supported_exts = {".csv", ".xlsx", ".xls", ".pdf"}

    # Build case-insensitive lookup:  header_lower → original header string
    header_lower_map = {h.lower(): h for h in column_headers}

    # First pass: accumulate all matched paths per column
    all_matches: dict = {}   # column_header → [file_path, ...]
    unrecognized: list = []

    if not os.path.isdir(folder_path):
        return {}, {}, []

    for filename in sorted(os.listdir(folder_path)):
        ext = os.path.splitext(filename)[1].lower()
        if ext not in supported_exts:
            continue  # Skip non-supported files silently

        stem = os.path.splitext(filename)[0]   # filename without extension
        stem_lower = stem.lower()

        if stem_lower in header_lower_map:
            col_header = header_lower_map[stem_lower]
            file_path = os.path.join(folder_path, filename)
            all_matches.setdefault(col_header, []).append(file_path)
        else:
            unrecognized.append(os.path.join(folder_path, filename))

    # Split into unambiguous matches and conflicts
    matches: dict = {}
    conflicts: dict = {}
    for col_header, paths in all_matches.items():
        if len(paths) == 1:
            matches[col_header] = paths[0]
        else:
            conflicts[col_header] = paths

    return matches, conflicts, unrecognized


def load_advisory_file(file_path: str) -> tuple:
    """
    Loads an advisory reference file and returns (content, error_message).

    content:
      - dict {code: description}  if the file has structured Code/Description columns
      - str                        if the file has text but no structured columns
      - None                       if parsing completely failed

    error_message:
      - None on success
      - str explaining the problem on failure
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return _load_csv(file_path)
    elif ext in (".xlsx", ".xls"):
        return _load_excel(file_path)
    elif ext == ".pdf":
        return _load_pdf(file_path)
    else:
        return None, f"Unsupported file format: '{ext}'"


def build_advisory_context(matches: dict, loaded_data: dict) -> dict:
    """
    Builds the final advisory context dict from matched files and loaded data.

    Returns {column_header: content} for each column whose file loaded successfully.
    Columns whose files failed to load are absent from the result.
    """
    result = {}
    for col_header in matches:
        content = loaded_data.get(col_header)
        if content is not None:
            result[col_header] = content
    return result


# ---------------------------------------------------------------------------
# Private loaders
# ---------------------------------------------------------------------------

def _load_csv(file_path: str) -> tuple:
    """
    Reads a CSV advisory file and returns a (data, error) tuple.

    Detection strategy (in priority order):
    1. If the CSV has a 'Code' column (case-insensitive), uses it plus an optional
       'Description' column → returns {code: description} dict (original behaviour).
    2. Fallback: scans every column; any column where ≥ 70 % of non-empty values are
       pure integers is treated as the code column (last such column wins).
       All remaining columns are joined into a single description string per code,
       with multiple rows sharing the same code concatenated together.
    3. If neither heuristic finds a code column, returns raw file text.
    """
    try:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            # Build case-insensitive map of fieldname → actual key
            fieldnames_lower = {k.lower(): k for k in (reader.fieldnames or [])}
            code_key = fieldnames_lower.get("code")
            desc_key = fieldnames_lower.get("description")

            # --- Path 1: explicit 'Code' column found ---
            # Multiple rows may share the same code (e.g. many asset names all have
            # life=60 months).  Concatenate all descriptions so keyword searches in
            # suggest_asset_life() can match any row, not just the last one.
            if code_key is not None:
                result = {}
                for row in reader:
                    code = str(row.get(code_key, "")).strip()
                    desc = str(row.get(desc_key, "")).strip() if desc_key else ""
                    if not code:
                        continue
                    if code in result:
                        # Append additional descriptions for the same code value
                        if desc:
                            result[code] = result[code] + ". " + desc
                    else:
                        result[code] = desc

                if not result:
                    return None, "CSV has a 'Code' column but no data rows were found."
                return result, None

            # --- Path 2: no explicit 'Code' column — detect by content heuristic ---
            # Consume all data rows now (header was already read by DictReader).
            rows = list(reader)

            # Scan each column to find one whose values are ≥ 70 % pure integers.
            # We keep the *last* matching column so that, for tables like
            # "Category, Account_Range, Asset_Name, …, Asset_Life_Months", the
            # rightmost numeric column (the life months) is preferred.
            detected_code_key = None
            for col in (reader.fieldnames or []):
                col_vals = [
                    str(r.get(col, "")).strip()
                    for r in rows
                    if str(r.get(col, "")).strip()
                ]
                if not col_vals:
                    continue
                int_count = sum(1 for v in col_vals if re.fullmatch(r"\d+", v))
                if int_count / len(col_vals) >= 0.70:
                    detected_code_key = col  # keep scanning; use last match

            if detected_code_key is None:
                # No structured numeric column found — return raw text as fallback
                f.seek(0)
                raw = f.read()
                return (raw if raw.strip() else None), None

            # Build {code: concatenated_description} from all non-code columns.
            # Rows sharing the same code are concatenated with ". " separator so
            # the keyword search in suggest_asset_life() can match any row's text.
            non_code_cols = [
                c for c in (reader.fieldnames or []) if c != detected_code_key
            ]
            result = {}
            for row in rows:
                code = str(row.get(detected_code_key, "")).strip()
                if not re.fullmatch(r"\d+", code):
                    # Skip header-like or blank values in the code column
                    continue
                parts = [str(row.get(c, "")).strip() for c in non_code_cols]
                row_text = " ".join(p for p in parts if p)
                if not row_text:
                    continue
                # Concatenate multiple rows with the same code value
                if code in result:
                    result[code] = result[code] + ". " + row_text
                else:
                    result[code] = row_text

            return (result if result else None), None

    except Exception as exc:
        return None, f"Could not read CSV: {exc}"


def _load_excel(file_path: str) -> tuple:
    """
    Reads an Excel advisory file (.xlsx / .xls).

    If the first sheet has 'Code' and 'Description' headers (case-insensitive),
    returns a {code: description} dict.  Otherwise returns raw concatenated text.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0]

        # Read header row to find Code / Description columns
        header = {}
        for cell in ws[1]:
            if cell.value:
                header[str(cell.value).strip().lower()] = cell.column_letter

        code_col = header.get("code")
        desc_col = header.get("description")

        if code_col is None:
            # No 'Code' column — return raw text from all cells as fallback
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                row_text = "  ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    rows_text.append(row_text.strip())
            raw = "\n".join(rows_text)
            return (raw if raw.strip() else None), None

        result = {}
        for row in ws.iter_rows(min_row=2):
            row_dict = {cell.column_letter: cell for cell in row}
            code_cell = row_dict.get(code_col)
            if code_cell is None or code_cell.value is None:
                continue
            code = str(code_cell.value).strip()
            desc = ""
            if desc_col and row_dict.get(desc_col) and row_dict[desc_col].value:
                desc = str(row_dict[desc_col].value).strip()
            if code:
                result[code] = desc

        if not result:
            return None, "Excel file has a 'Code' column but no data rows were found."
        return result, None

    except Exception as exc:
        return None, f"Could not read Excel file: {exc}"


def _extract_table_as_code_dict(file_path: str) -> tuple:
    """
    Fallback PDF parser for table-structured files where the asset name comes
    first and the numeric code (e.g. month value) appears at the end of each row.

    Algorithm:
      1. Open the PDF with pdfplumber and call page.extract_tables() on each page.
      2. Detect the "code column": prefer a column header containing "code"
         (case-insensitive); otherwise scan each column for one where ≥ 70% of
         non-empty values are pure integers — use the *last* such column found.
      3. Collect all non-code text cells from the same row and join them into a
         description string.
      4. Group rows by code value so that multiple rows sharing the same code are
         concatenated — keyword matching then covers every asset name for that code.

    Returns ({code_str: description}, None) if at least one entry was found,
    or (None, reason_str) otherwise.
    """
    try:
        import pdfplumber  # noqa: PLC0415 — deferred import; pdfplumber is optional
    except ImportError:
        return None, "pdfplumber is not installed"

    try:
        all_tables = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                # extract_tables() returns a list of tables; each table is a list of rows;
                # each row is a list of cell strings (or None for empty cells).
                tables = page.extract_tables()
                if tables:
                    all_tables.extend(tables)

        if not all_tables:
            return None, "No tables found in PDF"

        result: dict = {}  # code_str → concatenated description text

        for table in all_tables:
            if not table or len(table) < 2:
                continue  # Need at least a header row and one data row

            header_row = [str(c).strip() if c is not None else "" for c in table[0]]
            data_rows = table[1:]

            # Step 1: Try to find a column whose header contains "code" (case-insensitive)
            code_col_idx = None
            for col_idx, header_cell in enumerate(header_row):
                if "code" in header_cell.lower():
                    code_col_idx = col_idx
                    break  # Use first matching header column

            # Step 2: If no header match, scan data columns for one where ≥70% of
            # non-empty values are pure integers; use the last such column found.
            if code_col_idx is None:
                num_cols = len(header_row)
                for col_idx in range(num_cols):
                    col_values = []
                    for row in data_rows:
                        if col_idx < len(row) and row[col_idx] is not None:
                            val = str(row[col_idx]).strip()
                            if val:
                                col_values.append(val)
                    if not col_values:
                        continue
                    int_count = sum(1 for v in col_values if re.fullmatch(r"\d+", v))
                    if int_count / len(col_values) >= 0.70:
                        code_col_idx = col_idx  # Keep scanning — use the *last* match

            if code_col_idx is None:
                continue  # Cannot identify a code column in this table

            # Step 3: For each data row, extract the code and join the remaining cells
            # into a description string.
            for row in data_rows:
                if code_col_idx >= len(row):
                    continue
                code_val = row[code_col_idx]
                if code_val is None:
                    continue
                code_str = str(code_val).strip()
                if not re.fullmatch(r"\d+", code_str):
                    continue  # Skip rows where the code cell is not a plain integer

                # Collect text from all non-code columns in this row
                desc_parts = []
                for col_idx, cell in enumerate(row):
                    if col_idx == code_col_idx:
                        continue  # Skip the code column itself
                    if cell is not None:
                        cell_text = str(cell).strip()
                        if cell_text:
                            desc_parts.append(cell_text)

                row_text = " ".join(desc_parts).strip()
                if not row_text:
                    continue

                # Concatenate multiple rows that share the same code
                if code_str in result:
                    result[code_str] = result[code_str] + ". " + row_text
                else:
                    result[code_str] = row_text

        if result:
            return result, None
        return None, "Table extraction found no integer-coded rows"

    except Exception as exc:
        return None, f"Table extraction failed: {exc}"


def _load_pdf(file_path: str) -> tuple:
    """
    Reads a text-based PDF advisory file using pdfplumber and the same regex
    patterns used by tools/extract_ferc_pdf.py.

    Returns a {code: description} dict if structured entries are found.
    Falls back to returning raw extracted text if no structured entries are found.
    Returns (None, error_message) if the PDF cannot be read at all.
    """
    try:
        import pdfplumber  # noqa: PLC0415 — deferred import; pdfplumber is optional
    except ImportError:
        return None, "pdfplumber is not installed. Run: pip install pdfplumber"

    try:
        accounts_tabular: dict = {}   # from Style A patterns (preferred)
        accounts_header: dict = {}    # from Style B patterns (fallback)

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                lines = text.splitlines()

                # Track a "pending code" in case a code appears alone on a line
                # and its description continues on the next line.
                pending_code = None

                for line in lines:
                    # --- Resolve any pending lone code from the previous line ---
                    if pending_code is not None:
                        stripped = line.strip()
                        if (stripped
                                and not CODE_ONLY_RE.match(line)
                                and not TABULAR_RE.match(line)):
                            desc = _clean_description(stripped)
                            if desc:
                                accounts_tabular.setdefault(pending_code, desc)
                        pending_code = None

                    # --- Style A: tabular "101   Description" ---
                    m = TABULAR_RE.match(line)
                    if m:
                        code = m.group(1).strip()
                        desc = _clean_description(m.group(2))
                        if desc:
                            accounts_tabular.setdefault(code, desc)
                        continue

                    # --- Lone code on a line (description may wrap to next line) ---
                    m2 = CODE_ONLY_RE.match(line)
                    if m2:
                        pending_code = m2.group(1).strip()
                        continue

                    # --- Style B: "Account 101. Description" ---
                    m3 = HEADER_RE.search(line)
                    if m3:
                        code = m3.group(1).strip()
                        desc = _clean_description(m3.group(2))
                        if desc and code not in accounts_header:
                            accounts_header[code] = desc

        # Merge: tabular entries take precedence over header-style entries
        merged: dict = {}
        for code, desc in accounts_header.items():
            merged[code] = desc
        for code, desc in accounts_tabular.items():
            merged[code] = desc  # overwrites header entry if both found

        if merged:
            return merged, None

        # No structured entries from regex — try table extraction as a second pass.
        # This handles PDFs where the asset name appears first and the code (month value)
        # appears at the end of the row (e.g. the Asset Life depreciation schedule table),
        # which the line-oriented regex patterns above cannot parse.
        table_result, _table_err = _extract_table_as_code_dict(file_path)
        if table_result:
            return table_result, None

        # No structured entries found — try returning raw extracted text
        raw_pages = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    raw_pages.append(page_text)

        raw_text = "\n".join(raw_pages)
        if raw_text.strip():
            return raw_text, None

        # PDF opened but produced no text at all — likely a scanned image
        return None, (
            "PDF appears to be image-based (no extractable text). "
            "Only text-based PDFs are supported."
        )

    except Exception as exc:
        return None, f"Could not read PDF: {exc}"
