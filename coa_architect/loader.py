"""
loader.py — Reads the Chart of Accounts Excel file into Python objects.

CoALoader handles:
  - Column auto-detection (handles minor header-name variations)
  - Parsing of the main accounts worksheet into Account dataclass instances
  - Loading all reference sheets (FERC codes, asset life, cash flow, etc.)
  - Optionally merging an external FERC reference CSV or Excel file
"""

import csv
import os
from typing import Optional

import openpyxl

from .models import Account, ReferenceData


# ---------------------------------------------------------------------------
# Column header synonyms — allows the loader to find columns even when the
# header text differs slightly from one version of the file to another.
# ---------------------------------------------------------------------------
COLUMN_SYNONYMS = {
    "account_id": [
        "account id", "acct id", "id", "account_id", "acctid",
    ],
    "company": [
        "company", "co", "company code", "comp",
    ],
    "business_unit": [
        "business unit", "bu", "business_unit", "businessunit", "dept",
    ],
    "bu_type": [
        "bu type", "bu_type", "butype", "type", "bs/is", "balance sheet/income statement",
    ],
    "account_number": [
        "account number", "account no", "acct no", "acct number", "account#",
        "gl account", "gl acct", "account_number", "accountnumber",
    ],
    "account_description": [
        "account description", "description", "account name", "acct description",
        "account_description", "accountdescription", "name",
    ],
    "posting_edit": [
        "posting edit", "posting_edit", "postingedit", "edit code", "post edit",
    ],
    "line_of_detail": [
        "line of detail", "line_of_detail", "lineofdetail", "level", "lod",
        "detail level", "hierarchy level",
    ],
    "ferc_code": [
        "ferc code", "ferc_code", "fercode", "ferc", "ferc acct",
    ],
    "asset_life": [
        "asset life", "asset_life", "assetlife", "useful life", "life (months)",
        "life months", "depreciable life",
    ],
    "book_tax_difference": [
        "book tax difference", "book/tax difference", "book_tax_difference",
        "book tax diff", "btd",
    ],
    "cash_flow_category": [
        "cash flow category", "cash_flow_category", "cashflowcategory",
        "cash flow", "cf category", "cf cat",
    ],
}

# Reference sheet names we look for in the workbook (case-insensitive)
REFERENCE_SHEET_NAMES = {
    "ferc_codes": ["ferc", "ferc codes", "ferc code", "ferc ref"],
    "asset_life_codes": ["asset life", "asset_life", "asset life codes", "useful life"],
    "cash_flow_codes": ["cash flow", "cash flow category", "cash flow codes", "cf"],
    "posting_edit_codes": ["posting edit", "posting edit codes", "posting_edit"],
    "book_tax_codes": ["book tax", "book/tax", "book tax difference", "btd"],
    "companies": ["company", "companies", "company codes"],
    "business_units": ["business unit", "business units", "bu", "departments"],
}


class CoALoader:
    """
    Reads and parses a Chart of Accounts Excel workbook into Python objects.
    """

    def load_workbook(self, file_path: str) -> openpyxl.Workbook:
        """
        Opens an Excel file and returns the openpyxl Workbook object.
        data_only=True means we read computed cell values, not formulas.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"CoA file not found: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        return wb

    def detect_column_mapping(self, worksheet) -> dict:
        """
        Reads the first row of the worksheet to find which column letter
        corresponds to each logical field (e.g. 'account_number' → 'E').

        Header matching is case-insensitive and uses COLUMN_SYNONYMS for
        robustness against minor naming variations.

        Returns a dict like: {'account_id': 'A', 'account_number': 'E', ...}
        Raises ValueError if any required column cannot be found.
        """
        # Read first row as {col_letter: header_text_lower}
        header_row = {}
        for cell in worksheet[1]:
            if cell.value is not None:
                col_letter = cell.column_letter
                header_text = str(cell.value).strip().lower()
                header_row[col_letter] = header_text

        # Reverse map: header_text → col_letter
        text_to_col = {v: k for k, v in header_row.items()}

        mapping = {}
        for field_name, synonyms in COLUMN_SYNONYMS.items():
            found = False
            for synonym in synonyms:
                if synonym in text_to_col:
                    mapping[field_name] = text_to_col[synonym]
                    found = True
                    break
            if not found:
                # Non-critical fields (asset_life, book_tax, cash_flow) may be absent
                optional_fields = {"asset_life", "book_tax_difference", "cash_flow_category"}
                if field_name in optional_fields:
                    mapping[field_name] = None  # Signal that column is absent
                else:
                    raise ValueError(
                        f"Required column '{field_name}' not found in worksheet "
                        f"'{worksheet.title}'. Known headers: {list(header_row.values())}"
                    )

        return mapping

    def _cell_value(self, row_dict: dict, col_letter: Optional[str], default=None):
        """
        Retrieves a cell value from a pre-built {col_letter: cell} dict.
        Returns the default if col_letter is None or the cell is empty.
        """
        if col_letter is None:
            return default
        cell = row_dict.get(col_letter)
        if cell is None or cell.value is None:
            return default
        return cell.value

    def parse_accounts_sheet(self, worksheet, column_mapping: dict) -> list:
        """
        Iterates over every data row in the accounts worksheet (skipping header)
        and builds a list of Account objects.

        Skips rows that have no account_number (e.g. blank separator rows).
        """
        accounts = []

        # Identify the account_number column for the skip check
        num_col = column_mapping.get("account_number")

        for row in worksheet.iter_rows(min_row=2):
            # Build a quick lookup: column_letter → cell object
            row_dict = {cell.column_letter: cell for cell in row}

            # Skip rows without an account number (blank spacer rows)
            raw_num = self._cell_value(row_dict, num_col)
            if raw_num is None:
                continue

            # Parse account number — strip whitespace and convert to int
            try:
                account_number = int(str(raw_num).strip())
            except (ValueError, TypeError):
                continue  # Skip non-numeric rows (e.g. notes)

            # Parse line_of_detail — default to 5 if missing
            raw_lod = self._cell_value(row_dict, column_mapping.get("line_of_detail"))
            try:
                line_of_detail = int(str(raw_lod).strip()) if raw_lod is not None else 5
            except (ValueError, TypeError):
                line_of_detail = 5

            # Parse account_id — default to 0 if missing
            raw_id = self._cell_value(row_dict, column_mapping.get("account_id"), 0)
            try:
                account_id = int(str(raw_id).strip()) if raw_id is not None else 0
            except (ValueError, TypeError):
                account_id = 0

            # String fields — strip whitespace, convert None to ""
            def str_val(col_key, fallback=""):
                v = self._cell_value(row_dict, column_mapping.get(col_key))
                return str(v).strip() if v is not None else fallback

            account = Account(
                account_id=account_id,
                company=str_val("company"),
                business_unit=str_val("business_unit"),
                bu_type=str_val("bu_type"),
                account_number=account_number,
                account_description=str_val("account_description"),
                posting_edit=str_val("posting_edit"),
                line_of_detail=line_of_detail,
                ferc_code=str_val("ferc_code"),
                asset_life=str_val("asset_life") or None,
                book_tax_difference=str_val("book_tax_difference") or None,
                cash_flow_category=str_val("cash_flow_category") or None,
            )
            accounts.append(account)

        return accounts

    def parse_reference_sheet_as_lookup(
        self,
        worksheet,
        code_col: str = "A",
        desc_col: str = "B",
        skip_rows: int = 1,
    ) -> dict:
        """
        Reads a reference/lookup sheet and returns {code_str: description_str}.

        skip_rows: how many header rows to skip before data starts (default 1).
        code_col / desc_col: column letters for the code and description.
        """
        result = {}
        for row in worksheet.iter_rows(min_row=skip_rows + 1):
            row_dict = {cell.column_letter: cell for cell in row}

            code_cell = row_dict.get(code_col)
            desc_cell = row_dict.get(desc_col)

            if code_cell is None or code_cell.value is None:
                continue

            code = str(code_cell.value).strip()
            desc = str(desc_cell.value).strip() if (desc_cell and desc_cell.value) else ""
            if code:
                result[code] = desc

        return result

    def _parse_business_unit_sheet(self, worksheet) -> dict:
        """
        Parses the Business Unit reference sheet, which uses a non-standard layout:
        a variable-length header section is followed by a table whose columns are
        identified by a header row containing 'Business Unit Number'.

        Scans each row looking for the header row (any cell whose text contains
        'business unit number', case-insensitive).  Once found, records which
        column holds the BU number and which holds the description, then reads
        all subsequent data rows.

        Returns {bu_number_str: description_str}.
        Falls back to the generic parser (col A = code, col B = desc) if the
        expected header row is not found.
        """
        bu_col_idx = None    # 0-based index of the BU number column
        desc_col_idx = None  # 0-based index of the description column
        data_min_row = None  # 1-based row number where data begins

        # --- Pass 1: find the header row ---
        for row in worksheet.iter_rows():
            for i, cell in enumerate(row):
                cell_text = str(cell.value).strip().lower() if cell.value else ""
                if "business unit number" in cell_text or cell_text in ("bu number", "bu code"):
                    bu_col_idx = i
                if "description" in cell_text and bu_col_idx is not None:
                    desc_col_idx = i

            if bu_col_idx is not None:
                # Data starts on the row after this header row
                data_min_row = row[0].row + 1
                break

        # Fallback if no recognised header is found
        if bu_col_idx is None:
            return self.parse_reference_sheet_as_lookup(worksheet, "A", "B", skip_rows=1)

        # --- Pass 2: read data rows ---
        result = {}
        for row in worksheet.iter_rows(min_row=data_min_row):
            cells = list(row)
            if bu_col_idx >= len(cells):
                continue

            bu_cell = cells[bu_col_idx]
            if bu_cell.value is None:
                continue

            bu_code = str(bu_cell.value).strip()
            if not bu_code:
                continue

            desc = ""
            if desc_col_idx is not None and desc_col_idx < len(cells):
                desc_cell = cells[desc_col_idx]
                desc = str(desc_cell.value).strip() if desc_cell.value else ""

            result[bu_code] = desc

        return result

    def _find_sheet(self, workbook: openpyxl.Workbook, name_candidates: list):
        """
        Finds a worksheet by checking each name in name_candidates
        against the actual sheet names (case-insensitive).
        Returns the worksheet or None if not found.
        """
        sheet_names_lower = {s.lower(): s for s in workbook.sheetnames}
        for candidate in name_candidates:
            actual = sheet_names_lower.get(candidate.lower())
            if actual:
                return workbook[actual]
        return None

    def load_all_reference_data(self, workbook: openpyxl.Workbook) -> ReferenceData:
        """
        Loads all reference sheets from the workbook into a ReferenceData object.
        Missing sheets produce empty dicts — the app continues without them.
        """
        def load_ref(name_key: str) -> dict:
            sheet = self._find_sheet(workbook, REFERENCE_SHEET_NAMES[name_key])
            if sheet is None:
                return {}
            # Most reference sheets have the code in column A, description in column B
            # and a single header row
            return self.parse_reference_sheet_as_lookup(sheet, "A", "B", skip_rows=1)

        return ReferenceData(
            ferc_codes=load_ref("ferc_codes"),
            asset_life_codes=load_ref("asset_life_codes"),
            cash_flow_codes=load_ref("cash_flow_codes"),
            posting_edit_codes=load_ref("posting_edit_codes"),
            book_tax_codes=load_ref("book_tax_codes"),
            companies=load_ref("companies"),
            business_units=self._parse_business_unit_sheet(
                self._find_sheet(workbook, REFERENCE_SHEET_NAMES["business_units"])
            ) if self._find_sheet(workbook, REFERENCE_SHEET_NAMES["business_units"]) else {},
        )

    def load_chart_of_accounts(self, file_path: str):
        """
        Main entry point for loading a CoA file.

        Returns (accounts_list, reference_data, column_mapping, workbook)
        where:
          - accounts_list: list of Account objects from the main accounts sheet
          - reference_data: ReferenceData loaded from all reference sheets
          - column_mapping: dict of field_name → column_letter
          - workbook: the open openpyxl.Workbook (kept open for the exporter)
        """
        workbook = self.load_workbook(file_path)

        # The main accounts sheet is assumed to be the first sheet in the workbook
        main_sheet = workbook.worksheets[0]

        column_mapping = self.detect_column_mapping(main_sheet)
        accounts = self.parse_accounts_sheet(main_sheet, column_mapping)
        reference_data = self.load_all_reference_data(workbook)

        return accounts, reference_data, column_mapping, workbook

    def load_external_ferc_file(self, file_path: str) -> dict:
        """
        Loads additional FERC codes from a user-supplied CSV or Excel file.

        The file must have columns named 'Code' and 'Description' (case-insensitive).
        Returns {code_str: description_str}.
        This dict is merged into ReferenceData.ferc_codes by the caller,
        with CoA-embedded codes taking precedence over external ones.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"External FERC file not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            return self._load_ferc_from_csv(file_path)
        elif ext in (".xlsx", ".xls", ".xlsm"):
            return self._load_ferc_from_excel(file_path)
        else:
            raise ValueError(
                f"Unsupported external FERC file format: '{ext}'. "
                "Expected .csv or .xlsx"
            )

    def _load_ferc_from_csv(self, file_path: str) -> dict:
        """Reads FERC codes from a CSV file with 'Code' and 'Description' columns."""
        result = {}
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            # Normalize header names to lowercase for matching
            fieldnames_lower = {k.lower(): k for k in (reader.fieldnames or [])}
            code_key = fieldnames_lower.get("code")
            desc_key = fieldnames_lower.get("description")

            if code_key is None:
                raise ValueError(
                    f"External FERC CSV '{file_path}' must have a 'Code' column."
                )

            for row in reader:
                code = str(row.get(code_key, "")).strip()
                desc = str(row.get(desc_key, "")).strip() if desc_key else ""
                if code:
                    result[code] = desc

        return result

    def _load_ferc_from_excel(self, file_path: str) -> dict:
        """Reads FERC codes from an Excel file with 'Code' and 'Description' columns."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0]

        # Find Code and Description columns from the header row
        header = {}
        for cell in ws[1]:
            if cell.value:
                header[str(cell.value).strip().lower()] = cell.column_letter

        code_col = header.get("code")
        desc_col = header.get("description")

        if code_col is None:
            raise ValueError(
                f"External FERC Excel '{file_path}' must have a 'Code' column."
            )

        result = {}
        for row in ws.iter_rows(min_row=2):
            row_dict = {cell.column_letter: cell for cell in row}
            code_cell = row_dict.get(code_col)
            if code_cell is None or code_cell.value is None:
                continue
            code = str(code_cell.value).strip()
            desc = ""
            if desc_col and row_dict.get(desc_col) and row_dict[desc_col].value:
                desc = str(row_dict[desc_col].value).strip()
            if code:
                result[code] = desc

        return result
