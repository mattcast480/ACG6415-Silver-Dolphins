"""
exporter.py — Writes the updated Chart of Accounts back to Excel.

CoAExporter handles:
  - Finding the correct row position for the new account (after last sibling)
  - Inserting a new row while preserving formatting
  - Assigning the new account a max+1 Account ID (col A)
  - Creating a timestamped backup before saving
  - Returning the saved file path for confirmation display

This module is the most technically fragile because openpyxl row insertion
can affect formulas, merged cells, and named ranges.  We keep it simple:
  - Insert rows using worksheet.insert_rows()
  - Copy cell values from the adjacent row for styling cues
  - Write each field using column_mapping for portability
"""

import os
import shutil
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from .models import Account, AccountHierarchy, NewAccountProposal


class CoAExporter:
    """
    Writes a new account to the Chart of Accounts Excel file.
    """

    def insert_account_row(
        self,
        workbook: openpyxl.Workbook,
        proposal: NewAccountProposal,
        hierarchy: AccountHierarchy,
    ) -> int:
        """
        Finds the correct insertion position and writes the new account row.

        Position logic:
          - Find all children of the proposed parent.
          - Sort them by account_number.
          - Insert immediately after the last child that has an account_number
            <= proposal.account_number (so the list stays sorted).
          - If no children exist, insert immediately after the parent row.

        Returns the row number where the new account was inserted.
        """
        worksheet = workbook.worksheets[0]
        col_map = hierarchy.column_mapping

        # Build a mapping of account_number → row_number in the worksheet
        num_col = col_map.get("account_number")
        if num_col is None:
            raise ValueError("Cannot find account_number column in column_mapping.")

        # Read all rows to find position
        row_positions = {}  # account_number → row_number
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            row_dict = {cell.column_letter: cell for cell in row}
            cell = row_dict.get(num_col)
            if cell and cell.value is not None:
                try:
                    acct_num = int(str(cell.value).strip())
                    row_positions[acct_num] = row_idx
                except (ValueError, TypeError):
                    pass

        # Determine insertion row: after the largest sibling number
        # that is still less than the proposed number (keeps list sorted)
        parent = proposal.suggested_parent
        insertion_row = None

        if parent and parent.children:
            # Find siblings whose numbers are <= proposed number
            preceding_siblings = sorted(
                [c.account_number for c in parent.children
                 if c.account_number <= proposal.account_number],
                reverse=True,
            )
            if preceding_siblings:
                last_preceding = preceding_siblings[0]
                insertion_row = row_positions.get(last_preceding)

        if insertion_row is None:
            # No preceding siblings: insert after the parent row itself
            if parent and parent.account_number in row_positions:
                insertion_row = row_positions[parent.account_number]
            else:
                # Last resort: append at the end of the data
                insertion_row = worksheet.max_row

        # Insert a blank row after the determined position
        insert_at = insertion_row + 1
        worksheet.insert_rows(insert_at)

        # Copy cell format from the row above for visual consistency
        self._copy_row_format(worksheet, insertion_row, insert_at, col_map)

        # Write the new account data into the inserted row
        self._write_account_row(worksheet, insert_at, proposal, col_map, hierarchy)

        return insert_at

    def _copy_row_format(
        self,
        worksheet,
        source_row: int,
        target_row: int,
        col_map: dict,
    ) -> None:
        """
        Copies cell formatting (font, fill, alignment, border) from the source
        row to the target row.  This gives the new row the same visual style
        as its neighbors without embedding hard-coded formatting.
        """
        for source_cell in worksheet[source_row]:
            target_cell = worksheet.cell(row=target_row, column=source_cell.column)
            # Copy each formatting attribute safely (some cells may not have styles)
            if source_cell.has_style:
                target_cell.font = source_cell.font.copy() if source_cell.font else Font()
                target_cell.fill = source_cell.fill.copy() if source_cell.fill else PatternFill()
                target_cell.alignment = (
                    source_cell.alignment.copy() if source_cell.alignment else Alignment()
                )
                target_cell.number_format = source_cell.number_format

    def _write_account_row(
        self,
        worksheet,
        row_number: int,
        proposal: NewAccountProposal,
        col_map: dict,
        hierarchy: AccountHierarchy,
    ) -> None:
        """
        Writes all fields of the proposal into the correct columns of the
        given row number, using the column_mapping for portability.
        """
        # New account_id = one more than the current maximum
        new_account_id = hierarchy.max_account_id + 1

        # Map field names to their values
        field_values = {
            "account_id": new_account_id,
            "company": proposal.company or "",
            "business_unit": proposal.business_unit or "",
            "bu_type": proposal.bu_type or "",
            "account_number": proposal.account_number,
            "account_description": proposal.account_description or "",
            "posting_edit": proposal.posting_edit or "",
            "line_of_detail": proposal.line_of_detail or 5,
            "ferc_code": proposal.ferc_code or "",
            "asset_life": proposal.asset_life or "",
            "book_tax_difference": proposal.book_tax_difference or "",
            "cash_flow_category": proposal.cash_flow_category or "",
        }

        # Write each field to the corresponding column letter
        for field_name, value in field_values.items():
            col_letter = col_map.get(field_name)
            if col_letter is None:
                continue  # Column not present in this workbook
            cell = worksheet[f"{col_letter}{row_number}"]
            cell.value = value

    def update_account_id_sequence(
        self,
        worksheet,
        col_map: dict,
    ) -> None:
        """
        Re-sequences the Account ID column (col A or wherever it maps) so that
        all account IDs are sequential integers starting from 1 after insertion.

        This corrects the shift caused by inserting a new row.
        """
        id_col = col_map.get("account_id")
        if id_col is None:
            return  # No ID column to update

        sequence = 1
        for row in worksheet.iter_rows(min_row=2):
            row_dict = {cell.column_letter: cell for cell in row}
            id_cell = row_dict.get(id_col)
            if id_cell is None:
                continue
            # Check if this row has any data (skip genuinely blank rows)
            has_data = any(
                c.value is not None for c in row if c.column_letter != id_col
            )
            if has_data:
                id_cell.value = sequence
                sequence += 1

    def save_workbook(
        self,
        workbook: openpyxl.Workbook,
        file_path: str,
        create_backup: bool = True,
    ) -> str:
        """
        Saves the workbook to file_path, creating a timestamped backup first.

        Backup name format: Silver_Dolphins_CoA_backup_20260303_142255.xlsx
        Returns the path to the saved file.
        """
        if create_backup and os.path.exists(file_path):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(file_path)
            backup_path = f"{base}_backup_{timestamp}{ext}"
            shutil.copy2(file_path, backup_path)
            print(f"Backup saved: {backup_path}")

        workbook.save(file_path)
        return file_path

    def export(
        self,
        proposal: NewAccountProposal,
        hierarchy: AccountHierarchy,
        workbook: openpyxl.Workbook,
    ) -> str:
        """
        Main export entry point.

        Inserts the new account row (assigning max+1 Account ID) and saves the file.
        Returns the path to the saved file.
        """
        # Insert the row into the workbook
        self.insert_account_row(workbook, proposal, hierarchy)

        # Save with backup
        saved_path = self.save_workbook(workbook, hierarchy.source_file_path)
        return saved_path
