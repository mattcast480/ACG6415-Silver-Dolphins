"""
test_loader.py — Unit tests for coa_architect.loader.CoALoader.

Tests verify:
  - Correct number of accounts loaded from the CoA workbook
  - Column E is mapped to 'account_number'
  - Reference sheets are loaded (ferc_codes, asset_life_codes, etc.)
  - External FERC file loading (CSV and Excel formats)
  - Graceful handling of missing optional columns

Run with: pytest tests/test_loader.py -v
"""

import os
import tempfile
import pytest
import openpyxl

from coa_architect.loader import CoALoader
from coa_architect.models import Account, ReferenceData


# ---------------------------------------------------------------------------
# Fixtures — build minimal in-memory workbooks for testing
# ---------------------------------------------------------------------------

def _make_minimal_coa_workbook():
    """
    Creates a minimal Chart of Accounts workbook with:
      - A main accounts sheet (headers + 5 data rows)
      - A FERC codes reference sheet
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"

    # Header row matching the expected column names
    headers = [
        "Account ID",        # A
        "Company",           # B
        "Business Unit",     # C
        "BU Type",           # D
        "Account Number",    # E
        "Account Description",  # F
        "Posting Edit",      # G
        "Line of Detail",    # H
        "FERC Code",         # I
        "Asset Life",        # J
        "Book Tax Difference",  # K
        "Cash Flow Category",   # L
    ]
    ws.append(headers)

    # Sample data rows
    rows = [
        # id, co, bu, type, acct#, desc, pe, lod, ferc, life, btd, cf
        [1, "10", "10", "BS", 100000, "ASSETS", "", 1, "", None, None, None],
        [2, "10", "10", "BS", 100001, "Long-Term Assets", "", 2, "", None, None, None],
        [3, "10", "10", "BS", 100002, "Property Plant and Equipment", "", 3, "", None, None, None],
        [4, "10", "10", "BS", 100003, "Land", "", 4, "310", None, None, None],
        [5, "10", "10", "BS", 100500, "Land — Tract A", "", 5, "310", "0", None, None],
    ]
    for row in rows:
        ws.append(row)

    # FERC codes reference sheet
    ferc_ws = wb.create_sheet("FERC")
    ferc_ws.append(["Code", "Description"])
    ferc_ws.append(["310", "Land and land rights"])
    ferc_ws.append(["314", "Turbogenerator units"])

    return wb


@pytest.fixture
def minimal_coa_file(tmp_path):
    """Saves the minimal CoA workbook to a temp file and returns the path."""
    wb = _make_minimal_coa_workbook()
    path = tmp_path / "test_coa.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.fixture
def loader():
    return CoALoader()


# ---------------------------------------------------------------------------
# Tests — Column Detection
# ---------------------------------------------------------------------------

class TestDetectColumnMapping:
    def test_standard_headers_detected(self, minimal_coa_file, loader):
        """Column E should map to 'account_number'."""
        wb = loader.load_workbook(minimal_coa_file)
        ws = wb.worksheets[0]
        mapping = loader.detect_column_mapping(ws)

        assert mapping["account_number"] == "E", (
            f"Expected account_number → E, got {mapping['account_number']}"
        )

    def test_all_required_fields_present(self, minimal_coa_file, loader):
        """All required column names should be detected."""
        wb = loader.load_workbook(minimal_coa_file)
        ws = wb.worksheets[0]
        mapping = loader.detect_column_mapping(ws)

        required = [
            "account_id", "company", "business_unit", "bu_type",
            "account_number", "account_description", "posting_edit",
            "line_of_detail", "ferc_code",
        ]
        for field in required:
            assert field in mapping, f"Missing required field: {field}"
            assert mapping[field] is not None, f"Field {field} has None column letter"


# ---------------------------------------------------------------------------
# Tests — Account Parsing
# ---------------------------------------------------------------------------

class TestParseAccountsSheet:
    def test_correct_number_of_accounts(self, minimal_coa_file, loader):
        """Should load exactly 5 accounts (one per data row)."""
        wb = loader.load_workbook(minimal_coa_file)
        ws = wb.worksheets[0]
        mapping = loader.detect_column_mapping(ws)
        accounts = loader.parse_accounts_sheet(ws, mapping)

        assert len(accounts) == 5, f"Expected 5 accounts, got {len(accounts)}"

    def test_account_types(self, minimal_coa_file, loader):
        """All returned items should be Account instances."""
        wb = loader.load_workbook(minimal_coa_file)
        ws = wb.worksheets[0]
        mapping = loader.detect_column_mapping(ws)
        accounts = loader.parse_accounts_sheet(ws, mapping)

        for acct in accounts:
            assert isinstance(acct, Account)

    def test_account_numbers_parsed_correctly(self, minimal_coa_file, loader):
        """Account numbers should match the values in the workbook."""
        wb = loader.load_workbook(minimal_coa_file)
        ws = wb.worksheets[0]
        mapping = loader.detect_column_mapping(ws)
        accounts = loader.parse_accounts_sheet(ws, mapping)

        numbers = {a.account_number for a in accounts}
        expected = {100000, 100001, 100002, 100003, 100500}
        assert numbers == expected, f"Account numbers mismatch: {numbers} vs {expected}"

    def test_line_of_detail_parsed(self, minimal_coa_file, loader):
        """Line of detail values should be parsed as integers."""
        wb = loader.load_workbook(minimal_coa_file)
        ws = wb.worksheets[0]
        mapping = loader.detect_column_mapping(ws)
        accounts = loader.parse_accounts_sheet(ws, mapping)

        level5 = [a for a in accounts if a.account_number == 100500]
        assert len(level5) == 1
        assert level5[0].line_of_detail == 5


# ---------------------------------------------------------------------------
# Tests — Reference Data Loading
# ---------------------------------------------------------------------------

class TestLoadReferenceData:
    def test_ferc_codes_loaded(self, minimal_coa_file, loader):
        """FERC codes should be loaded from the FERC sheet."""
        wb = loader.load_workbook(minimal_coa_file)
        ref = loader.load_all_reference_data(wb)

        assert isinstance(ref, ReferenceData)
        assert "310" in ref.ferc_codes, f"Expected FERC code 310, got {ref.ferc_codes}"

    def test_missing_sheets_return_empty_dicts(self, minimal_coa_file, loader):
        """Missing reference sheets should produce empty dicts, not errors."""
        wb = loader.load_workbook(minimal_coa_file)
        ref = loader.load_all_reference_data(wb)

        # These sheets weren't added to our test workbook
        assert isinstance(ref.asset_life_codes, dict)
        assert isinstance(ref.cash_flow_codes, dict)
        assert isinstance(ref.companies, dict)


# ---------------------------------------------------------------------------
# Tests — Full Load Pipeline
# ---------------------------------------------------------------------------

class TestLoadChartOfAccounts:
    def test_returns_all_expected_values(self, minimal_coa_file, loader):
        """load_chart_of_accounts should return (accounts, ref_data, col_map, workbook)."""
        result = loader.load_chart_of_accounts(minimal_coa_file)
        assert len(result) == 4, "Expected (accounts, reference_data, column_mapping, workbook)"

        accounts, ref_data, col_map, wb = result
        assert len(accounts) > 0
        assert isinstance(ref_data, ReferenceData)
        assert isinstance(col_map, dict)
        assert wb is not None

    def test_file_not_found_raises(self, loader):
        """Loading a non-existent file should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            loader.load_chart_of_accounts("/nonexistent/path/file.xlsx")


# ---------------------------------------------------------------------------
# Tests — External FERC File
# ---------------------------------------------------------------------------

class TestLoadExternalFercFile:
    def test_load_from_csv(self, tmp_path, loader):
        """Should load FERC codes from a CSV file."""
        csv_file = tmp_path / "ferc.csv"
        csv_file.write_text("Code,Description\n999,Test Code\n998,Another Code\n")

        result = loader.load_external_ferc_file(str(csv_file))
        assert "999" in result
        assert result["999"] == "Test Code"

    def test_load_from_excel(self, tmp_path, loader):
        """Should load FERC codes from an Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Code", "Description"])
        ws.append(["888", "Excel Code"])
        path = tmp_path / "ferc.xlsx"
        wb.save(str(path))

        result = loader.load_external_ferc_file(str(path))
        assert "888" in result
        assert result["888"] == "Excel Code"

    def test_missing_code_column_raises(self, tmp_path, loader):
        """CSV without a 'Code' column should raise ValueError."""
        csv_file = tmp_path / "bad.csv"
        csv_file.write_text("Number,Label\n100,Test\n")

        with pytest.raises(ValueError, match="Code"):
            loader.load_external_ferc_file(str(csv_file))

    def test_unsupported_format_raises(self, tmp_path, loader):
        """Unsupported file formats should raise ValueError."""
        txt_file = tmp_path / "ferc.txt"
        txt_file.write_text("100,Test")

        with pytest.raises(ValueError, match="Unsupported"):
            loader.load_external_ferc_file(str(txt_file))
